import matplotlib.pyplot as plt
import numpy as np
import csv
from operator import itemgetter
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00, FORMAT_NUMBER_00
from jinja2 import Environment, FileSystemLoader
import pdfkit

currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055, }


def take_ten_items(dictionary):
    """Берёт 10 первых пар словаря, округляя значения до 4 знаков после запятой
    Args:
        dictionary (dict): словарь
    Returns:
        dict: обрезанный словарь
    >>> take_ten_items({"a": 1.45609, "b": 6.666, "c": 7, "d": 0.52723, "e": 5, "f": 4,"g": 3, "h": 2, "i": 12.4444, "j": 11, "k": 12.4433, "l": 13.3, "m": 42.094394})
    {'a': 1.4561, 'b': 6.666, 'c': 7, 'd': 0.5272, 'e': 5, 'f': 4, 'g': 3, 'h': 2, 'i': 12.4444, 'j': 11}
    >>> take_ten_items({'a': 2.3699, 'b': 4.666, 'c': 5.898989, 'd': 0.7677, 'e': 5})
    {'a': 2.3699, 'b': 4.666, 'c': 5.899, 'd': 0.7677, 'e': 5}
    """
    new_dictionary = {}
    i = 0
    for key in dictionary:
        new_dictionary[key] = round(dictionary[key], 4)
        i += 1
        if i == 10:
            break
    return new_dictionary


class DataSet:
    """Составляет базу данных для вакансий.

    Attributes:
        file_name (string): название файла
        profession (string): название профессии
        vacancies_objects (list): список вакансий
        vacancies_count_by_years (dict): Словарь типа {ключ-год : значение-количество вакансий}
        vacancies_count_by_years_for_profession (dict): Словарь типа {ключ-год : значение-количество вакансий определённой профессии}
        salary_by_years (dict): Словарь типа {ключ-год : значение-уровень зарплат}
        salary_by_years_for_profession (dict): Словарь типа {ключ-год : значение-уровень зарплат определённой профессии}
        vacancies_count_by_cities (dict): Словарь типа {ключ-город : значение-количество вакансий}
        vacancies_share_by_cities (dict): Словарь типа {ключ-город : значение-процент вакансий от общего кол-ва}
        salary_by_cities (dict): Словарь типа {ключ-город : значение-уровень зарплат}
    """

    def __init__(self, file_name, profession):
        """Инициализируект объект DataSet, создаёт различные словари
        Args:
            file_name (string): название файла
            profession (string): профессия
        >>> type(DataSet("vacancies.csv", "Программист")).__name__
        'DataSet'
        >>> DataSet("vacancies.csv", "Программист").file_name
        'vacancies.csv'
        >>> DataSet("vacancies.csv", "Программист").profession
        'Программист'
        """
        self.file_name = file_name
        self.profession = profession
        headlines, vacancies = self.csv_reader()
        dictionaries = self.csv_filer(vacancies, headlines)
        self.vacancies_objects = [Vacancy(dictionary) for dictionary in dictionaries]
        self.vacancies_count_by_years = self.get_vacancies_count_by_years()
        self.vacancies_count_by_years_for_profession = self.get_vacancies_count_by_years_for_profession()
        self.salary_by_years = self.get_salary_by_years()
        self.salary_by_years_for_profession = self.get_salary_by_years_for_profession()
        self.vacancies_count_by_cities = self.get_vacancies_count_by_cities()
        self.vacancies_share_by_cities = self.get_vacancies_share_by_cities()
        self.salary_by_cities = self.get_salary_by_cities()

    def csv_reader(self):
        """Создаёт список вакансий и список их параметров
        Returns:
            list: список параметров вакансий
            list: список вакансий
        """
        headlines_list = []
        vacancies_list = []
        length = 0
        first_element = True
        rows_count = 0
        with open(self.file_name, encoding="utf-8-sig") as File:
            reader = csv.reader(File)
            for row in reader:
                rows_count += 1
                if first_element:
                    headlines_list = row
                    length = len(row)
                    first_element = False
                else:
                    need_to_break = False
                    if length != len(row):
                        need_to_break = True
                    for word in row:
                        if word == "":
                            need_to_break = True
                    if need_to_break:
                        continue
                    vacancies_list.append(row)
        if rows_count == 0:
            print("Пустой файл")
            exit()
        if rows_count == 1:
            print("Нет данных")
            exit()
        return headlines_list, vacancies_list

    def csv_filer(self, reader, list_naming):
        """Создаёт словарь вакансий и их параметров
        Args:
            reader (list): список вакансий
            list_naming (list): список параметров вакансий
        Returns:
            dict: словарь типа {вакансия : параметры}
        """
        dictionaries_list = []
        for vacancy in reader:
            dictionary = {}
            for i in range(len(list_naming)):
                dictionary[list_naming[i]] = vacancy[i]
            dictionaries_list.append(dictionary)
        return dictionaries_list

    def get_vacancies_count_by_years(self):
        """
        Возвращает словарь годов и кол-ва вакансий
        Returns:
            dict: Словарь типа {ключ-год : значение-количество вакансий}
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if vacancy.published_at in dictionary:
                dictionary[vacancy.published_at] += 1
            else:
                dictionary[vacancy.published_at] = 1
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary

    def get_vacancies_count_by_years_for_profession(self):
        """
        Возвращает словарь годов и кол-ва вакансий определённой профессии
        Returns:
            dict: Словарь типа {ключ-год : значение-количество вакансий определённой профессии}
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            if vacancy.published_at in dictionary:
                dictionary[vacancy.published_at] += 1
            else:
                dictionary[vacancy.published_at] = 1
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary = {2022: 0}
        return dictionary

    def get_salary_by_years(self):
        """
        Возвращает словарь годов и уровня зарплат
        Returns:
            dict: Словарь типа {ключ-год : значение-уровень зарплат}
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if vacancy.published_at in dictionary:
                dictionary[vacancy.published_at] += vacancy.salary
            else:
                dictionary[vacancy.published_at] = vacancy.salary
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        return dictionary

    def get_salary_by_years_for_profession(self):
        """
        Возвращает словарь годов и уровня зарплат определённой профессии
        Returns:
            dict: Словарь типа {ключ-год : значение-уровень зарплат определённой профессии}
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            if vacancy.published_at in dictionary:
                dictionary[vacancy.published_at] += vacancy.salary
            else:
                dictionary[vacancy.published_at] = vacancy.salary
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_years_for_profession[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(0)))
        if len(dictionary) == 0:
            dictionary = {2022: 0}
        return dictionary

    def get_vacancies_count_by_cities(self):
        """
        Возвращает словарь городов и кол-ва вакансий
        Returns:
            dict: Словарь типа {ключ-город : значение-кол-во вакансий}
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if vacancy.area_name in dictionary:
                dictionary[vacancy.area_name] += 1
            else:
                dictionary[vacancy.area_name] = 1
        return dictionary

    def get_vacancies_share_by_cities(self):
        """
        Возвращает словарь городов и процента вакансий от общего кол-ва
        Returns:
            dict: Словарь типа {ключ-город : значение-процент вакансий от общего кол-ва}
        """
        dictionary = {}
        for key in self.vacancies_count_by_cities:
            if self.vacancies_count_by_cities[key] / len(self.vacancies_objects) >= 0.01:
                dictionary[key] = self.vacancies_count_by_cities[key] / len(self.vacancies_objects)
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True))
        new_dictionary = take_ten_items(dictionary)
        return new_dictionary

    def get_salary_by_cities(self):
        """
        Возвращает словарь городов и уровня зарплат
        Returns:
            dict: Словарь типа {ключ-город : значение-уровень зарплат}
        """
        dictionary = {}
        for vacancy in self.vacancies_objects:
            if self.vacancies_count_by_cities[vacancy.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            if vacancy.area_name in dictionary:
                dictionary[vacancy.area_name] += vacancy.salary
            else:
                dictionary[vacancy.area_name] = vacancy.salary
        for key in dictionary:
            dictionary[key] = int(dictionary[key] / self.vacancies_count_by_cities[key])
        dictionary = dict(sorted(dictionary.items(), key=itemgetter(1), reverse=True))
        new_dictionary = take_ten_items(dictionary)
        return new_dictionary

    def print_information(self):
        """
        Выводит в консоль статистику вакансий
        """
        print("Динамика уровня зарплат по годам: " + str(self.salary_by_years))
        print("Динамика количества вакансий по годам: " + str(self.vacancies_count_by_years))
        print("Динамика уровня зарплат по годам для выбранной профессии: " + str(
            self.salary_by_years_for_profession))
        print("Динамика количества вакансий по годам для выбранной профессии: " + str(
            self.vacancies_count_by_years_for_profession))
        print("Уровень зарплат по городам (в порядке убывания): " + str(self.salary_by_cities))
        print("Доля вакансий по городам (в порядке убывания): " + str(self.vacancies_share_by_cities))


class Vacancy:
    """Класс для вакансии

    Attributes:
        name (string): название
        salary (string): зарплата
        area_name (string): город
        published_at (int): дата публикации
    """

    def __init__(self, dictionary):
        """Инициализируект объект Vacancy
        Args:
            dictionary (dict): словарь
        >>> type(Vacancy({"name": "Программист", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Сургут",  "published_at": "2020-07-05T18:19:30+0300"})).__name__
        'Vacancy'
        >>> Vacancy({"name": "Программист", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Сургут",  "published_at": "2020-07-05T18:19:30+0300"}).name
        'Программист'
        >>> Vacancy({"name": "Программист", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Сургут",  "published_at": "2020-07-05T18:19:30+0300"}).area_name
        'Сургут'
        >>> Vacancy({"name": "Программист", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Сургут",  "published_at": "2020-07-05T18:19:30+0300"}).salary
        55000.0
        >>> Vacancy({"name": "Программист", "salary_from": 10000, "salary_to": 100000, "salary_currency": "RUR", "area_name": "Сургут",  "published_at": "2020-07-05T18:19:30+0300"}).published_at
        2020
        """
        self.name = dictionary["name"]
        self.salary = (float(dictionary["salary_from"]) + float(dictionary["salary_to"])) / 2 * currency_to_rub[
            dictionary["salary_currency"]]
        self.area_name = dictionary["area_name"]
        self.published_at = int(dictionary["published_at"][:4])


class Report:
    """Класс для визуализации статистики

    Attributes:
        profession (string): профессия
        years_list_headers (list): список заголовков, связынных с годами
        years_list_columns (list): список годов и параметров, связанных с ними
        cities_list_headers (list): список заголовков, связынных с городами
        cities_list_columns (list): список городов и параметров, связанных с ними
        years_list_widths (list): список ширин для таблицы по годам
        cities_list_widths (list): список ширин для таблицы по городам
    """

    def __init__(self, dataset):
        """Инициализируект объект Report, формирует различные данные
        Args:
            dat1aset (DataSet): dataset
        """
        self.profession = dataset.profession
        self.years_list_headers = (
            "Год", "Средняя зарплата", f"Средняя зарплата - {profession}", "Количество вакансий",
            f"Количество вакансий - {profession}")
        self.years_list_columns = [[year for year in dataset.salary_by_years],
                                   [value for value in dataset.salary_by_years.values()],
                                   [value for value in dataset.salary_by_years_for_profession.values()],
                                   [value for value in dataset.vacancies_count_by_years.values()],
                                   [value for value in dataset.vacancies_count_by_years_for_profession.values()]]

        self.cities_list_headers = ("Город", "Уровень зарплат", "", "Город", "Доля вакансий")
        self.cities_list_columns = [[city for city in dataset.salary_by_cities],
                                    [value for value in dataset.salary_by_cities.values()],
                                    ["" for i in range(len(dataset.salary_by_cities))],
                                    [city for city in dataset.vacancies_share_by_cities],
                                    [value for value in dataset.vacancies_share_by_cities.values()]]

        self.years_list_widths = [len(header) + 2 for header in self.years_list_headers]
        for i in range(len(self.years_list_columns)):
            for cell in self.years_list_columns[i]:
                self.years_list_widths[i] = max(len(str(cell)) + 2, self.years_list_widths[i])

        self.cities_list_widths = [len(header) + 2 for header in self.cities_list_headers]
        for i in range(len(self.cities_list_columns)):
            for cell in self.cities_list_columns[i]:
                self.cities_list_widths[i] = max(len(str(cell)) + 2, self.cities_list_widths[i])

    def set_border(self, ws, width, height):
        """Устанавливает рамки для таблицы
        Args:
            ws (openpyxl.Workbook()): Excel лист
            width (float): ширина
            height (float) высота
        """
        cell_range = f'A1:{get_column_letter(width)}{height}'
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def clear_column(self, ws, column):
        """Удаляет рамки для столбца таблицы
        Args:
            ws (openpyxl.Workbook()): Excel лист
            column (string): буква, соответствующая столбцу
        """
        empty = Side(border_style=None)
        for cell in ws[column]:
            cell.border = Border(top=empty, bottom=empty)

    def get_vertical_chart(self, title, parameter_1, parameter_1_name, parameter_2,
                           parameter_2_name, labels, fig, number):
        """Создаёт вертикальную диаграмму
        Args:
            title (string): Название
            parameter_1 (list): первый параметр
            parameter_1_name (string): название первого параметра
            parameter_2 (list): второй параметр
            parameter_2_name (string): название втрого параметра
            labels (list): лейблы
            fig (plt.figure()): фигура
            number (int): расположение графика на листе
        """
        x = np.arange(len(labels))
        width = 0.35
        plt.rcParams['font.size'] = '8'
        ax = fig.add_subplot(number)
        ax.bar(x - width / 2, parameter_1, width, label=parameter_1_name)
        ax.bar(x + width / 2, parameter_2, width, label=parameter_2_name)
        ax.set_xticks(x, labels, rotation="vertical")
        ax.grid(axis='y')
        ax.set_title(title)
        ax.legend()

    def get_horizontal_chart(self, title, parameter, labels, fig):
        """Создаёт горизонтальную диаграмму
        Args:
            title (string): Название
            parameter (list): параметр
            labels (list): лейблы
            fig (plt.figure()): фигура
        """
        plt.rcParams['font.size'] = '8'
        ax = fig.add_subplot(223)
        labels = [city.replace(' ', '\n').replace('-', '-\n') for city in labels]
        y = np.arange(len(labels))
        ax.barh(y, parameter)
        ax.set_yticks(y, labels=labels, fontsize=6)
        ax.grid(axis='x')
        ax.invert_yaxis()
        ax.set_title(title)

    def get_pie_chart(self, title, parameter, labels, fig):
        """Создаёт вертикальную диаграмму
        Args:
            title (string): Название
            parameter (list): параметр
            labels (list): лейблы
            fig (plt.figure()): фигура
        """
        plt.rcParams['font.size'] = '6'
        labels.insert(0, "Другие")
        parameter.insert(0, 1 - sum(parameter))
        ax = fig.add_subplot(224)
        ax.pie(parameter, labels=labels)
        ax.axis('equal')
        ax.set_title(title)
        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_excel(self):
        """Создаёт вертикальную диаграмму
        Returns:
            years_list (wb): лист для таблицы по годам
            cities_list (wb): лист для таблицы по городам
        """
        wb = openpyxl.Workbook()
        years_list = wb.active
        years_list.title = "Статистика по годам"
        cities_list = wb.create_sheet("Статистика по городам")
        years_list.append(self.years_list_headers)
        for cell in years_list['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.years_list_columns[0])):
            years_list.append([column[i] for column in self.years_list_columns])
        cities_list.append(self.cities_list_headers)
        for cell in cities_list['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.cities_list_columns[0])):
            cities_list.append([column[i] for column in self.cities_list_columns])
        for cell in cities_list['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        for i in range(1, 6):
            years_list.column_dimensions[get_column_letter(i)].width = self.years_list_widths[i - 1]
            cities_list.column_dimensions[get_column_letter(i)].width = self.cities_list_widths[i - 1]
        self.set_border(years_list, len(self.years_list_headers), len(self.years_list_columns[0]) + 1)
        self.set_border(cities_list, len(self.cities_list_headers), len(self.cities_list_columns[0]) + 1)
        self.clear_column(cities_list, 'C')
        wb.save('report.xlsx')
        return years_list, cities_list

    def generate_image(self):
        """Генерирует изображение"""
        fig = plt.figure()
        self.get_vertical_chart("Уровень зарплат по годам", self.years_list_columns[1], "средняя з/п",
                                self.years_list_columns[2], f"з/п {self.profession}", self.years_list_columns[0], fig,
                                221)
        self.get_vertical_chart("Количество вакансий по годам", self.years_list_columns[3], "Количество вакансий",
                                self.years_list_columns[4], f"Количество вакансий {self.profession}",
                                self.years_list_columns[0], fig, 222)
        self.get_horizontal_chart("Уровень зарплат по городам", self.cities_list_columns[1],
                                  self.cities_list_columns[0], fig)
        self.get_pie_chart("Доля вакансий по городам", self.cities_list_columns[4], self.cities_list_columns[3], fig)
        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self):
        """Генерирует pdf документ"""
        years_list, cities_list = self.generate_excel()
        self.generate_image()
        for cell in cities_list['E']:
            if cell.value == "Доля вакансий":
                continue
            cell.value = str(round(float(cell.value) * 100, 2)).replace('.', ',') + '%'

        env = Environment(loader=FileSystemLoader('.'))
        pdf_template = env.get_template("pdf_template.html").render(
            {'profession': f'{self.profession}', 'image_file': "graph.png",
             'years_list': years_list, 'cities_list': cities_list})
        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={'enable-local-file-access': None})


def run_program():
    """Запускает программу"""
    global profession
    file_name = input("Введите название файла: ")
    profession = input("Введите профессию: ")
    dataset = DataSet(file_name, profession)
    dataset.print_information()
    Report(dataset).generate_pdf()