import csv
from operator import itemgetter
import openpyxl
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00


class DataSet:
    def __init__(self, name, profession):
        self.file_name = name
        self.profession = profession
        headlines, vacancies = self.csv_reader()
        dictionaries = self.csv_filter(vacancies, headlines)
        self.vacancies_objects = [Vacancy(dictionary) for dictionary in dictionaries]
        self.vacancies_count_by_years = self.get_vacancies_count_by_years()
        self.vacancies_count_by_years_for_profession = self.get_vacancies_count_by_years_for_profession()
        self.salary_by_years = self.get_salary_by_years()
        self.salary_by_years_for_profession = self.get_salary_by_years_for_profession()
        self.vacancies_count_by_cities = self.get_vacancies_count_by_cities()
        self.vacancies_share_by_cities = self.get_vacancies_share_by_cities()
        self.salary_by_cities = self.get_salary_by_cities()

    def csv_reader(self):
        length = 0
        count = 0
        headlines_list = []
        vacancies_list = []
        is_first = True
        with open(name, encoding="utf-8-sig") as File:
            reader = csv.reader(File)
            for row in reader:
                count += 1
                if is_first:
                    headlines_list = row
                    length = len(row)
                    is_first = False
                else:
                    need_to_continue = False
                    if length != len(row):
                        need_to_continue = True
                    for word in row:
                        if word == "":
                            need_to_continue = True
                    if need_to_continue:
                        continue
                    vacancies_list.append(row)
        if count == 0:
            print("Здесь написано третье другое")
            exit()
        elif count == 1:
            print("Нет данных")
            exit()
        return headlines_list, vacancies_list

    def csv_filter(self, reader, list_naming):
        dict_list = []
        for vacancy in reader:
            dictionary = {}
            for i in range(len(list_naming)):
                dictionary[list_naming[i]] = vacancy[i]
            dict_list.append(dictionary)
        return dict_list

    def dict_sotrer(self, dict):
        return dict(sorted(dict.items(), key=itemgetter(0)))

    def dict_processing_published(self, dict, vacancy, summand):
        if vacancy.published_at in dict:
            dict[vacancy.published_at] += summand
        else:
            dict[vacancy.published_at] = summand
        return dict

    def dict_processing_area(self, dict, vacancy, summand):
        if vacancy.area_name in dict:
            dict[vacancy.area_name] += summand
        else:
            dict[vacancy.area_name] = summand
        return dict

    def get_vacancies_count_by_years(self):
        dict = {}
        for vacancy in self.vacancies_objects:
            dict = self.dict_processing_published(dict, vacancy, 1)
        dict = self.dict_sotrer(dict)
        return dict

    def get_vacancies_count_by_years_for_profession(self):
        dict = {}
        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dict = self.dict_processing_published(dict, vacancy, 1)
        dict = self.dict_sotrer(dict)
        if len(dict) == 0:
            dict = {2022: 0}
        return dict

    def get_salary_by_years(self):
        dict = {}
        for vacancy in self.vacancies_objects:
            dict = self.dict_processing_published(dict, vacancy, vacancy.salary)
        for key in dict:
            dict[key] = int(dict[key] / self.vacancies_count_by_years[key])
        dict = self.dict_sotrer(dict)
        return dict

    def get_salary_by_years_for_profession(self):
        dict = {}

        for vacancy in self.vacancies_objects:
            if self.profession not in vacancy.name:
                continue
            dict = self.dict_processing_published(dict, vacancy, vacancy.salary)
        for key in dict:
            dict[key] = int(dict[key] / self.vacancies_count_by_years_for_profession[key])
        dict = self.dict_sotrer(dict)
        if len(dict) == 0:
            dict = {2022: 0}
        return dict

    def get_vacancies_count_by_cities(self):
        dict = {}
        for vacancy in self.vacancies_objects:
            dict = self.dict_processing_area(dict, vacancy, 1)
        return dict

    def get_vacancies_share_by_cities(self):
        dict = {}
        for key in self.vacancies_count_by_cities:
            if self.vacancies_count_by_cities[key] / len(self.vacancies_objects) >= 0.01:
                dict[key] = self.vacancies_count_by_cities[key] / len(self.vacancies_objects)
        dict = dict(sorted(dict.items(), key=itemgetter(1), reverse=True))
        new_dict = self.take_ten_items(dict)
        return new_dict

    def get_salary_by_cities(self):
        dict = {}
        for vacancy in self.vacancies_objects:
            if self.vacancies_count_by_cities[vacancy.area_name] / len(self.vacancies_objects) < 0.01:
                continue
            dict = self.dict_processing_area(dict, vacancy, vacancy.salary)
        for key in dict:
            dict[key] = int(dict[key] / self.vacancies_count_by_cities[key])
        dict = dict(sorted(dict.items(), key=itemgetter(1), reverse=True))
        new_dict = self.take_ten_items(dict)
        return new_dict

    def print_information(self):
        print(f"Динамика уровня зарплат по годам: {str(self.salary_by_years)}")
        print(f"Динамика количества вакансий по годам: {str(self.vacancies_count_by_years)}")
        print(f"Динамика уровня зарплат по годам для выбранной профессии: {str(self.salary_by_years_for_profession)}")
        print(f"Динамика количества вакансий по годам для выбранной профессии: {str(self.vacancies_count_by_years_for_profession)}")
        print(f"Уровень зарплат по городам (в порядке убывания): {str(self.salary_by_cities)}")
        print(f"Доля вакансий по городам (в порядке убывания): {str(self.vacancies_share_by_cities)}")

    def take_ten_items(self, dictionary):
        new_dict = {}
        i = 0
        for key in dictionary:
            new_dict[key] = round(dictionary[key], 4)
            i += 1
            if i == 10:
                break
        return new_dict

class Vacancy:
    def __init__(self, dict):
        self.name = dict["name"]
        self.salary = (float(dict["salary_from"]) + float(dict["salary_to"])) / 2 * currency[
            dict["salary_currency"]]
        self.area_name = dict["area_name"]
        self.published_at = int(dict["published_at"][:4])


class Report:
    def __init__(self, dataset):
        self.list_top_headers = (
            "Год", "Средняя зарплата", f"Средняя зарплата - {dataset.profession}", "Количество вакансий",
            f"Количество вакансий - {dataset.profession}")
        self.list_top_columns = [[year for year in dataset.salary_by_years],
                                 [value for value in dataset.salary_by_years.values()],
                                 [value for value in dataset.salary_by_years_for_profession.values()],
                                 [value for value in dataset.vacancies_count_by_years.values()],
                                 [value for value in dataset.vacancies_count_by_years_for_profession.values()]]

        self.cities_list_headers = ("Город", "Уровень зарплат", "", "Город", "Доля вакансий")
        self.cities_list_columns = [[city for city in dataset.salary_by_cities],
                                    [value for value in dataset.salary_by_cities.values()],
                                    ["" for i in range(len(dataset.salary_by_cities))],
                                    [city for city in dataset.vacancies_share_by_cities],
                                    [value for value in dataset.vacancies_share_by_cities.values()]]

        self.years_list_widths = [len(header) + 2 for header in self.list_top_headers]
        for i in range(len(self.list_top_columns)):
            for cell in self.list_top_columns[i]:
                self.years_list_widths[i] = max(len(str(cell)) + 2, self.years_list_widths[i])

        self.cities_list_widths = [len(header) + 2 for header in self.cities_list_headers]
        for i in range(len(self.cities_list_columns)):
            for cell in self.cities_list_columns[i]:
                self.cities_list_widths[i] = max(len(str(cell)) + 2, self.cities_list_widths[i])

    def set_borders(self, ws, width, height):
        cell_range = f'A1:{get_column_letter(width)}{height}'
        thin = Side(border_style="thin", color="000000")
        for row in ws[cell_range]:
            for cell in row:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    def clear_column(self, ws, column):
        empty = Side(border_style=None)
        for cell in ws[column]:
            cell.border = Border(top=empty, bottom=empty)

    def generate_excel(self):
        work_book = openpyxl.Workbook()
        years_list = work_book.active
        years_list.title = "Статистика по годам"
        cities_list = work_book.create_sheet("Статистика по городам")
        years_list.append(self.list_top_headers)
        for cell in years_list['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.list_top_columns[0])):
            years_list.append([column[i] for column in self.list_top_columns])
        cities_list.append(self.cities_list_headers)
        for cell in cities_list['1']:
            cell.font = Font(bold=True)
        for i in range(len(self.cities_list_columns[0])):
            cities_list.append([column[i] for column in self.cities_list_columns])
        for cell in cities_list['E']:
            cell.number_format = FORMAT_PERCENTAGE_00
        for i in range(1, 6):
            years_list.column_dimensions[get_column_letter(i)].width = self.years_list_widths[i - 1]
            cities_list.column_dimensions[get_column_letter(i)].width = self.cities_list_widths[i - 1]
        self.set_borders(years_list, len(self.list_top_headers), len(self.list_top_columns[0]) + 1)
        self.set_borders(cities_list, len(self.cities_list_headers), len(self.cities_list_columns[0]) + 1)
        self.clear_column(cities_list, 'C')
        work_book.save('report.xlsx')


currency = {"AZN": 35.68,
            "BYR": 23.91,
            "EUR": 59.90,
            "GEL": 21.74,
            "KGS": 0.76,
            "KZT": 0.13,
            "RUR": 1,
            "UAH": 1.64,
            "USD": 60.66,
            "UZS": 0.0055, }

name = input("Введите название файла: ")
profession = input("Введите название профессии: ")
dataset = DataSet(name, profession)
dataset.print_information()
Report(dataset).generate_excel()
